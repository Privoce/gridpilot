"""Phase 8 — Engineering report and export generators.

Documents added to the packet on top of the model files:
  - Equipment schedule (.xlsx)
  - Load flow & short-circuit validation report (.pdf)
  - Dynamic model validation report with bump plots (.pdf)
  - Assumptions & approvals log (.pdf)
  - Source trace appendix (.md)
  - Portal-ready field export (.json)

All content reads from the engineering pass (`run_engineering`) — nothing is
re-derived here, so the reports cannot drift from the models.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import fitz


def _pdf_helpers():
    """Deferred import to avoid a circular module load with caiso_packet."""
    from backend.app.services.caiso_packet import (
        _Pdf, _axes, _polyline, ACC, INK, MUT, OKC, RED,
    )
    return _Pdf, _axes, _polyline, ACC, INK, MUT, OKC, RED


# ---------------------------------------------------------------------------
# Equipment schedule (.xlsx)
# ---------------------------------------------------------------------------

def gen_equipment_schedule(eng: dict[str, Any], intake: dict[str, Any], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Schedule"
    head_fill = PatternFill("solid", fgColor="1A1D1B")
    head_font = Font(color="FFFFFF", bold=True, size=10)

    ws["A1"] = f"{intake.get('project_name')} — Equipment Schedule"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"Generated from project graph "
                f"{eng['graph'].get('version', '')} · topology: {eng['design']['topology']}")
    ws["A2"].font = Font(size=8, color="666666")

    cols = ["Item", "Make / model", "Qty", "Rating", "Data source", "OEM verified"]
    widths = [26, 34, 8, 34, 40, 14]
    for i, (c, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=4, column=i, value=c)
        cell.fill = head_fill
        cell.font = head_font
        ws.column_dimensions[cell.column_letter].width = w

    r = 5
    for row in eng["design"]["schedule"]:
        ws.cell(row=r, column=1, value=row["item"])
        ws.cell(row=r, column=2, value=row["make_model"])
        ws.cell(row=r, column=3, value=row["qty"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value=row["rating"])
        ws.cell(row=r, column=5, value=row["source"])
        ws.cell(row=r, column=6, value="Yes" if row["verified"] else "CONFIRM")
        if not row["verified"]:
            ws.cell(row=r, column=6).font = Font(color="B45309", bold=True)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Counts").font = Font(bold=True)
    r += 1
    for k, v in eng["design"]["counts"].items():
        ws.cell(row=r, column=1, value=k.replace("_", " ").title())
        ws.cell(row=r, column=2, value=v)
        r += 1
    wb.save(path)


# ---------------------------------------------------------------------------
# Load flow + short circuit validation report (.pdf)
# ---------------------------------------------------------------------------

def gen_loadflow_report(eng: dict[str, Any], intake: dict[str, Any], path: Path) -> None:
    _Pdf, _axes, _polyline, ACC, INK, MUT, OKC, RED = _pdf_helpers()
    pf, sc = eng["powerflow"], eng["short_circuit"]
    pdf = _Pdf("Load Flow & Short-Circuit Validation",
               f"{intake.get('project_name')} — solved plant equivalent, "
               f"{pf['p_poi_mw']:g} MW at the POI",
               banner="COMPUTED — deterministic engineering engine")

    pdf.section("Solution summary")
    pdf.kv("Method", "Backward/forward sweep (radial plant equivalent)")
    pdf.kv("Convergence", f"{pf['iterations']} iterations to {pf['tolerance']:.0e} pu")
    pdf.kv("Dispatch", f"PV {pf['p_pv_mw']:g} MW + BESS {pf['p_bess_mw']:g} MW "
                       f"− aux {pf['aux_mw']:g} MW")
    pdf.kv("Delivered at POI", f"{pf['p_poi_mw']:g} MW / {pf['q_poi_mvar']:g} Mvar",
           f"requested {pf['p_requested_mw']:g} MW — "
           + ("match" if pf["poi_match"] else "MISMATCH"))
    pdf.kv("Series losses (computed)", f"{pf['losses_mw']:g} MW / {pf['losses_mvar']:g} Mvar",
           f"declared in intake: {pf['losses_declared_mw']:g} MW (Δ {pf['loss_delta_mw']:+g})")
    tap_pct = pf.get("mpt_tap_pct", 0.0)
    pdf.kv("MPT tap (OLTC)", f"{pf.get('mpt_tap', 1.0):.4f} ({tap_pct:+g}%)",
           "adjusted to hold the collector bus within 0.98-1.02 pu" if tap_pct
           else "nominal — collector bus within 0.98-1.02 pu")
    base_case = intake.get("file_base_case")
    if isinstance(base_case, dict) and base_case.get("name") and not base_case.get("staged"):
        pdf.kv("System representation", f"Customer base case on file: {base_case['name']}",
               "run the final validation against it in the authorized environment")
    else:
        pdf.kv("System representation", "Flat system equivalent at the POI",
               "no customer base case supplied — ISO base cases are confidential and are "
               "never fabricated by GridPilot")

    pdf.section("Bus voltage profile")
    for v in pf["voltages"]:
        pdf.kv(v["bus"], f"{v['v_pu']:.4f} pu ∠ {v['angle_deg']:g}°")

    pdf.section("Branch flows and loadings")
    for f in pf["flows"]:
        loading = f"{f['loading_pct']:g}% of {f['rating_mva']:g} MVA" if f["loading_pct"] is not None else "—"
        pdf.kv(f["branch"], f"{f['mva']:g} MVA · {loading}",
               f"losses {f['loss_mw']:g} MW / {f['loss_mvar']:g} Mvar")

    pdf.section("Convergence history")
    pdf.para("Max complex voltage mismatch per sweep iteration: "
             + ", ".join(pf["history"]), size=8, color=MUT)
    pdf.section("Dispatch iterations (POI power match)")
    for it in pf["dispatch"]:
        pdf.kv(f"Iteration {it['iter']}",
               f"gen {it['p_gen_mw']:g} MW → POI {it['p_poi_mw']:g} MW")

    pdf.section("Short-circuit duty (three-phase bolted)")
    pdf.para(sc["assumption"], size=8.5, color=MUT)
    for rrow in sc["results"]:
        pdf.kv(rrow["location"], f"{rrow['duty_mva']:g} MVA · {rrow['duty_ka']:g} kA",
               f"suggested breaker class: {rrow['breaker_class']}")
    pdf.kv("Plant contribution", f"{sc['plant_contribution_mva']:g} MVA",
           "inverter fleet as current-limited sources")

    if pf["warnings"]:
        pdf.section("Warnings")
        for w in pf["warnings"]:
            pdf.bullet(w, color=RED)

    pdf.section("Scope & source boundary")
    pdf.para("Computed by GridPilot's deterministic engine from the source-traced project "
             "graph. This validates the plant-side design; the official study model must be "
             "run against the ISO base case (obtained under NDA) in the ISO-designated "
             "software before submission.", size=8.5, color=MUT)
    pdf.save(path)


# ---------------------------------------------------------------------------
# Dynamic validation report with bump plots (.pdf)
# ---------------------------------------------------------------------------

def _plot(page, rect, t, series, color, y_label, events, _axes, _polyline, MUT, ACC):
    _axes(page, rect, y_label, "seconds", "")
    lo, hi = min(series), max(series)
    pad = max((hi - lo) * 0.15, 1e-3)
    lo, hi = lo - pad, hi + pad
    t0, t1 = t[0], t[-1]

    def X(tv):
        return rect.x0 + (tv - t0) / (t1 - t0) * rect.width

    def Y(v):
        return rect.y1 - (v - lo) / (hi - lo) * rect.height

    pts = [(X(tv), Y(v)) for tv, v in zip(t, series)]
    _polyline(page, pts, color, 1.3)
    for ev in events:
        page.draw_line(fitz.Point(X(ev["t"]), rect.y0), fitz.Point(X(ev["t"]), rect.y1),
                       color=(0.85, 0.85, 0.85), width=0.7)
    page.insert_text((rect.x0 + 4, rect.y0 + 10), f"{max(series):.2f} max", fontsize=6.5,
                     fontname="helv", color=MUT)
    page.insert_text((rect.x0 + 4, rect.y1 - 4), f"{min(series):.2f} min", fontsize=6.5,
                     fontname="helv", color=MUT)


def gen_dynamics_report(eng: dict[str, Any], intake: dict[str, Any], path: Path) -> None:
    _Pdf, _axes, _polyline, ACC, INK, MUT, OKC, RED = _pdf_helpers()
    bump = eng["bump"]
    inv = eng["design"]["inverter"]
    pdf = _Pdf("Dynamic Model Validation — Flat Start & Bump Tests",
               f"{intake.get('project_name')} — "
               f"{inv['wecc_models']['gen']}/{inv['wecc_models']['elec']}/{inv['wecc_models']['plant']} "
               f"plant response",
               banner="COMPUTED — deterministic engineering engine")

    pdf.section("Validation checks")
    for c in bump["checks"]:
        pdf.checkbox(c["pass"], f"{c['name']} — {c['detail']}")

    pdf.section("Events")
    for ev in bump["events"]:
        pdf.kv(f"t = {ev['t']:g} s", ev["label"])

    # Plots page.
    pdf._new_page()
    page = pdf.page
    t = bump["t"]
    events = bump["events"]
    rects = [fitz.Rect(70, 120, 545, 250), fitz.Rect(70, 300, 545, 430),
             fitz.Rect(70, 480, 545, 610)]
    _plot(page, rects[0], t, bump["p_mw"], ACC, "Plant P at POI (MW)", events,
          _axes, _polyline, MUT, ACC)
    _plot(page, rects[1], t, bump["q_mvar"], OKC, "Plant Q at POI (Mvar)", events,
          _axes, _polyline, MUT, ACC)
    _plot(page, rects[2], t, bump["v_pu"], RED, "POI voltage (pu)", events,
          _axes, _polyline, MUT, ACC)
    pdf.y = 640

    pdf.section("Model basis & boundary")
    pdf.para(f"PV fleet: {eng['design']['counts']['inverters']} x {inv['vendor']} "
             f"{inv['model']} — parameters from {inv['datasheet']}.", size=8.5)
    bu = eng["design"].get("bess_unit")
    if bu and eng["design"]["counts"]["bess_units"]:
        pdf.para(f"BESS: {eng['design']['counts']['bess_units']} x {bu['vendor']} {bu['model']} — "
                 f"parameters from {bu['datasheet']}.", size=8.5)
    pdf.para(bump["model_note"], size=8.5, color=MUT)
    pdf.save(path)


# ---------------------------------------------------------------------------
# Assumptions & approvals log (.pdf)
# ---------------------------------------------------------------------------

def gen_assumptions_log(eng: dict[str, Any], intake: dict[str, Any], path: Path) -> None:
    _Pdf, _axes, _polyline, ACC, INK, MUT, OKC, RED = _pdf_helpers()
    ap = eng["approvals"]
    pdf = _Pdf("Assumptions & Approvals Log",
               f"{intake.get('project_name')} — engineering judgment items and sign-off state",
               banner="APPROVAL WORKFLOW")

    pdf.section("Summary")
    pdf.kv("Assumption items", str(ap["total"]))
    pdf.kv("Approved", str(ap["total"] - ap["pending"]))
    pdf.kv("Pending", str(ap["pending"]),
           "" if ap["all_approved"] else "Packet remains draft until all items are approved")

    pdf.section("Ledger")
    for item in ap["items"]:
        state = "APPROVED" if item["approved"] else "PENDING"
        who = ""
        if item["approved"] and isinstance(item["approval"], dict):
            who = f" by {item['approval'].get('by', '?')} on {item['approval'].get('at', '?')}"
        pdf.kv(item["label"], f"{item['value']} {item['unit']}".strip(),
               f"{state}{who} — {item.get('trace') or item['origin']}")

    if eng["graph"].get("missing"):
        pdf.section("Missing engineering inputs")
        for m in eng["graph"]["missing"]:
            pdf.bullet(f"{m['label']} — impacts: {m['impact']}", color=RED)

    pdf.section("Cross-document consistency checks")
    for c in eng["checks"]:
        pdf.checkbox(c["status"] == "pass", f"{c['name']} — {c['detail']}")

    pdf.section("Professional boundary")
    pdf.para("Items in this log are engineering defaults or derived values that require "
             "confirmation by the developer's engineer of record. Legal certifications and "
             "PE approvals are executed outside GridPilot and are not replaced by this log.",
             size=8.5, color=MUT)
    pdf.save(path)


# ---------------------------------------------------------------------------
# Source trace appendix (.md) and portal fields (.json)
# ---------------------------------------------------------------------------

def gen_source_trace(eng: dict[str, Any], intake: dict[str, Any], path: Path) -> None:
    g = eng["graph"]
    lines = [
        f"# Source Trace — {intake.get('project_name')}",
        "",
        f"Project graph `{g.get('version', '')}` · generated "
        f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} by the engineering engine.",
        "",
        "Every parameter used in this packet, with its source and origin. "
        "Source classes: developer, document_extraction, oem_datasheet, iso_rule, "
        "gridpilot_calculation, assumption_requiring_approval.",
        "",
        "| Parameter | Value | Source | Origin / formula |",
        "|---|---|---|---|",
    ]
    for r in eng["source_trace"]:
        val = f"{r['value']} {r['unit']}".strip()
        origin = r["trace"] or r["origin"]
        lines.append(f"| {r['label']} | {val} | {r['source_label']} | {origin} |")
    lines += [
        "",
        "## Topology",
        "",
        f"`{eng['design']['topology']}`",
        "",
        "## Solved case",
        "",
        f"- Converged in {eng['powerflow']['iterations']} iterations "
        f"(tolerance {eng['powerflow']['tolerance']:.0e})",
        f"- POI delivery: {eng['powerflow']['p_poi_mw']} MW / "
        f"{eng['powerflow']['q_poi_mvar']} Mvar",
        f"- Series losses: {eng['powerflow']['losses_mw']} MW",
    ]
    history = eng.get("graph_history") or []
    if history:
        lines += ["", "## Graph version history", "",
                  "| Version | When | Who | What | Changed |",
                  "|---|---|---|---|---|"]
        for h in history:
            changed = ", ".join(h.get("changed") or []) or "—"
            lines.append(
                f"| `{h.get('version', '')}` | {h.get('at', '')} | {h.get('by', '')} | "
                f"{h.get('note', '')} | {changed} |"
            )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def gen_missing_data(eng: dict[str, Any], intake: dict[str, Any], path: Path) -> None:
    """Missing-data list: facts the engine needed but the developer hasn't provided.

    The blueprint rule is that absent engineering truth becomes a tracked list
    item, never a guess — this document is that list, packaged for the
    developer alongside what the engine used in the meantime.
    """
    missing = eng["design"]["missing"]
    assume = {a["id"]: a for a in eng["approvals"]["items"]}
    lines = [
        f"# Missing Data List — {intake.get('project_name')}",
        "",
        f"Project graph `{eng['graph'].get('version', '')}` · generated "
        f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} by the engineering engine.",
        "",
        f"{len(missing)} item(s) the engineering engine needed but could not find in the "
        "intake or attached documents. GridPilot never invents missing engineering truth: "
        "each item below is either carried as an approvable assumption or left open, and the "
        "affected deliverables are listed so the impact is clear.",
        "",
        "| # | Missing item | Intake field | Affects |",
        "|---|---|---|---|",
    ]
    for i, m in enumerate(missing, 1):
        lines.append(f"| {i} | {m['label']} | `{m['key']}` | {m['impact']} |")
    if assume:
        lines += [
            "",
            "## Engineering defaults in use meanwhile",
            "",
            "These assumptions stand in for missing facts and require engineer sign-off "
            "(see the Assumptions & Approvals Log):",
            "",
        ]
        for a in assume.values():
            state = "approved" if a.get("approved") else "PENDING approval"
            lines.append(f"- **{a['label']}** = {a['value']} {a.get('unit', '')} — "
                         f"{a.get('trace') or a.get('origin') or ''} ({state})")
    lines += [
        "",
        "Provide the missing items and regenerate the packet — every affected document, "
        "model, and drawing updates from the same project graph.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def gen_portal_fields(eng: dict[str, Any], intake: dict[str, Any],
                      profile: dict[str, Any], path: Path) -> None:
    pf = eng["powerflow"]
    fields = {
        "portal": profile["portal"],
        "portal_url": profile.get("portal_url", ""),
        "generated_by": "deterministic engineering engine",
        "graph_version": eng["graph"].get("version", ""),
        "fields": {
            "applicant_legal_name": intake.get("legal_name"),
            "project_name": intake.get("project_name"),
            "authorized_signatory": intake.get("signatory_name"),
            "signatory_title": intake.get("signatory_title"),
            "contact_email": intake.get("contact_email"),
            "contact_phone": intake.get("contact_phone"),
            "project_type": intake.get("project_type"),
            "process_track": intake.get("track"),
            "deliverability": intake.get("deliverability"),
            "requested_cod": intake.get("cod"),
            "poi_name": intake.get("poi_name"),
            "poi_voltage_kv": intake.get("poi_voltage_kv"),
            "gentie_mi": eng["design"]["gentie_mi"],
            "shared_facilities": intake.get("shared_facilities"),
            "prior_queue_reference": intake.get("queue_ref"),
            "county_state": f"{intake.get('county')}, {intake.get('state')}",
            "gps": {"lat": intake.get("gps_lat"), "lon": intake.get("gps_lon")},
            "site_control": intake.get("site_control"),
            "site_acreage": intake.get("site_acreage"),
            "gross_mw": intake.get("gross_mw"),
            "aux_mw": intake.get("aux_mw"),
            "losses_mw_declared": intake.get("losses_mw"),
            "losses_mw_computed": pf["losses_mw"],
            "net_mw_at_poi": pf["p_poi_mw"],
            "bess_mw": intake.get("bess_mw"),
            "bess_mwh": intake.get("bess_mwh"),
            "inverter": intake.get("inverter"),
            "dynamic_models": eng["design"]["inverter"]["wecc_models"],
            "equipment_counts": eng["design"]["counts"],
        },
        "approvals": {
            "total": eng["approvals"]["total"],
            "pending": eng["approvals"]["pending"],
        },
        "note": "Copy these values into the portal fields; e-signature and payment "
                "remain with the authorized officer.",
    }
    path.write_text(json.dumps(fields, indent=2, default=str), encoding="utf-8")
